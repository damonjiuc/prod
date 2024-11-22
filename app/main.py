import openpyxl
import datetime
import mysql.connector


def transform_date(date_str):
    """ Дата вида dd.mm.yyyy в yyyy-mm-dd """
    if date_str is None:
        return None
    return '-'.join(reversed(date_str.split('.')))

def read_goods(goods_xlsx, orders_to_check):
    """
    Выбираем из выгрузки эксель те товыры, которые привязаны к заказам (ключ словаря orders_to_check).
    Добавляем в них id заказов к которым они привязаны.
    Убираем из БД все строки прявязанные к этим заказам.
    Возвращаем словарь со списками товаров привязанных к заказам.
    """
    if len(orders_to_check) == 0:
        return {}

    wb = openpyxl.load_workbook(goods_xlsx)
    sheet = wb.active
    goods_by_order = {}
    orders_ids = {}

    # alch_orders_ids = {}
    # for order in orders_to_check:


    orders_to_check_str = "', '".join(orders_to_check)
    sql = f"SELECT ORDER_ID, ORDER_NUM FROM orders WHERE ORDER_NUM IN ('{orders_to_check_str}');"
    cursor.execute(sql)
    db_orders_ids = cursor.fetchall()

    for order in db_orders_ids:
        orders_ids[order[1]] = order[0]

    for i in range(2, sheet.max_row + 1):
        order_num = sheet[f'A{i}'].value  # номер заказа
        if order_num not in orders_ids.keys():
            continue

        order_id = orders_ids[order_num] # id заказа
        cursor.execute(f'DELETE FROM items WHERE order_id ="{order_id}"; ') # Убираем из БД все товары привязанные к новым/изменненным заказам

        name = sheet[f'B{i}'].value  # имя
        amount = sheet[f'C{i}'].value  # количество
        amount_type = sheet[f'D{i}'].value  # шт/упаковки/..
        price = sheet[f'E{i}'].value  # Цена штуки

        if order_num in orders_to_check:
            cur_order_goods = goods_by_order.get(order_num, list())
            cur_order_goods.append([order_id, order_num, name, amount, amount_type, price])
            goods_by_order[order_num] = cur_order_goods

    return goods_by_order

def read_deals(orders_xlsx='xls/Dveri diz zakazy (XLSX).xlsx', days_to_load=32):
    """
    Считываем выгрузку заказов из 1с
    :return: словарь key - номер заказа, value - список значений
    """
    wb = openpyxl.load_workbook(orders_xlsx)
    sheet = wb.active
    selected_date = datetime.datetime.now().date() - datetime.timedelta(days_to_load)
    orders = {}


    for i in range(2, sheet.max_row + 1):
        order = sheet[f'A{i}'].value  # Заказ
        order_num = sheet[f'B{i}'].value  # Номер заказа
        order_date = sheet[f'C{i}'].value  # Дата заказа
        if datetime.datetime.strptime(order_date, '%d.%m.%Y').date() < selected_date:
            continue

        order_date = transform_date(order_date)
        shipment_date = transform_date(sheet[f'D{i}'].value)  # Дата отгрузки
        payment_date = transform_date(sheet[f'E{i}'].value)  # Дата выплаты
        bonus = sheet[f'F{i}'].value  # Бонус
        customer_name = sheet[f'G{i}'].value  # Покупатель
        order_sum = sheet[f'H{i}'].value  # Сумма заказа
        card_num = sheet[f'I{i}'].value  # Номер карты
        card_was_given = transform_date(sheet[f'J{i}'].value)  # Карта выдана
        address = sheet[f'K{i}'].value  # Адрес
        shop = sheet[f'L{i}'].value  # Магазин

        orders[order_num] = [order, order_num, order_date, shipment_date, payment_date, bonus, customer_name, order_sum, card_num, card_was_given, address, shop]

    return orders

def process_orders(orders_dict_xls):
    """
    Обработка заказов, апдейт и инсерт в БД

    :param orders_dict_xls: список (номер заказа: значения параметров заказов)
    :return: список заказов по которым нужно проверить товары
    """
    orders_to_check = list(orders_dict_xls.keys())
    orders_insert = {}
    orders_update = {}

    orders_dict_db = read_orders_from_db([f'"{order_num}"' for order_num in orders_to_check])

    for order in orders_to_check:
        if order in orders_dict_db.keys():
            if str(orders_dict_db[order][1]) == str(orders_dict_xls[order][2]) and str(orders_dict_db[order][2]) == str(orders_dict_xls[order][3]) and str(orders_dict_db[order][3]) == str(orders_dict_xls[order][4]) and orders_dict_db[order][5] == float(orders_dict_xls[order][7]):
                continue
            else:
                orders_update[order] = orders_dict_xls[order]
        else:
            orders_insert[order] = orders_dict_xls[order]

    for order in orders_insert.values():
        for i_el in range(len(order)):
            if order[i_el] is None:
                order[i_el] = ''

        cursor.execute("""INSERT INTO orders(ORDER_NAME, ORDER_NUM, ORDER_DATE, SHIPMENT_DATE, PAYMENT_DATE, BONUS, CUSTOMER_NAME, ORDER_SUM, CARD_NUM, CARD_WAS_GIVEN, ADDRESS, SHOP)
                    VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""", order)

    for order in orders_update.values():
        for i_el in range(len(order)):
            if order[i_el] is None:
                order[i_el] = ''

        cursor.execute(f"UPDATE orders SET ORDER_DATE='{order[2]}', SHIPMENT_DATE='{order[3]}', PAYMENT_DATE='{order[4]}', BONUS='{order[5]}', ORDER_SUM='{order[7]}' WHERE ORDER_NUM= '{order[1]}'")

    goods_insert = list(orders_insert.keys())
    goods_update = list(orders_update.keys())
    goods_to_check = goods_insert + goods_update

    print('insert', goods_insert)
    print('update', goods_update)
    print('check', goods_to_check)

    return goods_to_check


def read_orders_from_db(orders_nums):
    """
    Чтение заказов из БД вместе с присвоенным id auto increment

    :param orders_nums: номера заказов для обработки
    :return: словарь (номер заказа: параметры заказа)
    """
    orders_dict = {}
    cursor.execute(f"SELECT ORDER_ID, ORDER_NUM, ORDER_DATE, SHIPMENT_DATE, PAYMENT_DATE, BONUS, ORDER_SUM FROM orders WHERE ORDER_NUM IN ({', '.join(orders_nums)});")
    orders_tab =  cursor.fetchall()
    for order in orders_tab:
        orders_dict[order[1]] = [order[0], order[2], order[3], order[4], order[5], order[6]]

    return orders_dict

def process_one_direction(orders_table, items_table):
    """
    Обработка выгрузки по 1 направлению

    :param orders_table: линк на таблицу с заказами
    :param items_table: линк на таблицу с предметами из заказов
    :return:
    """
    print(f'Processin {orders_table}, {items_table}')

    cur_orders = read_deals(orders_table, 9999)

    if len(cur_orders.keys()) == 0:
        return None

    check_goods = process_orders(cur_orders)

    cur_goods = read_goods(items_table, check_goods)

    for order_goods in cur_goods.values():
        for good in order_goods:
            for i_el in range(len(good)):
                if good[i_el] is None:
                    good[i_el] = ''

            cursor.execute("""INSERT INTO items(ORDER_ID, ORDER_NUM, NAME, AMOUNT, AMOUNT_TYPE, PRICE)
                            VALUES(%s, %s, %s, %s, %s, %s)""", good)

    print()

def process_users_vars():
    """ Функция для расчета сумм, которые потратили дизайнеры и и уровня партнерки """
    # Получаем из БД инфо о юзерах и кладем в словарь, ключ - номер карты
    cursor.execute(f"SELECT ID, CARD, LVL, MONEYTOMLN, TOTALMONEY FROM users")
    db_users_w_money = cursor.fetchall()
    users_w_money_dict = {}
    for user in db_users_w_money:
        users_w_money_dict[str(user[1])] = user

    # Выгружаем из БД заказы, привязаные к юзерам
    sql_users_w_money = "', '".join(users_w_money_dict.keys())
    cursor.execute(
        f"SELECT ORDER_ID, ORDER_DATE, ORDER_SUM, CARD_NUM FROM orders WHERE CARD_NUM IN ('{sql_users_w_money}');")
    db_orders_w_sum = cursor.fetchall()

    # Создаем словарь, где ключ - номер карты юзера, а значение - список [Сумма заказов до 23г, Сумма заказов после 23г, Сумма за текущий год]
    # Проходим циклом по всем заказам и плюсуем сумму в нужную переменную
    users_orders = {}

    for order in db_orders_w_sum:
        cur_user_orders = users_orders.get(order[3], [0, 0, 0])
        if order[1] < datetime.date(2023, 1, 1):
            cur_user_orders[0] += int(order[2])
        else:
            cur_user_orders[1] += int(order[2])
            if order[1] > datetime.date(2024, 1, 1):
                cur_user_orders[2] += int(order[2])

        users_orders[order[3]] = cur_user_orders

    # Добавляем в список сумм каждого юзера общую сумму заказов (до и после 23г)
    for user in users_orders.keys():
        users_orders[user].append(users_orders[user][0] + users_orders[user][1])

    # Расчет и запись в БД
    for user, money in users_orders.items():
        total_money = money[-1]
        money_this_year = money[-2]
        lvl_a = int(money[0] // 500000)  # уровень до 23г
        money_after = money[1] + money[0] % 500000  # потрачено после 23г с учетом остатка до 23г
        lvl_b = int(money_after // 1000000)  # уровень накопленный после 23г
        lvl = lvl_a + lvl_b  # уровень суммарный
        if lvl > 10:  # проверка, что уровень не больше 10
            lvl = 10
        money_to_lvl = money_after % 1000000
        total_money_db = int(users_w_money_dict[str(user)][-1])
        lvl_db = int(users_w_money_dict[str(user)][2])
        tickets = int(money_this_year // 500000)
        print(
            f'\nUser {user}: уровней до 23 - {lvl_a}, уровней после 23 - {lvl_b}, \nвсего потрачено {total_money} / {money_this_year} в этом году, \nнакоплено для повышения уровня {money_to_lvl}')
        # если рассчитанные значения уровня и суммы заказов не такие, как в БД - перезаписываем БД
        if total_money != total_money_db or lvl != lvl_db:
            sql = f"UPDATE users SET LVL='{lvl}', MONEYTOMLN='{money_to_lvl}', TOTALMONEY='{total_money}', MONEYTHISYEAR='{money_this_year}', TICKETS='{tickets}' WHERE CARD='{user}'"
            cursor.execute(sql)

if __name__ == '__main__':
    db = mysql.connector.connect(
        host="vh436.timeweb.ru",
        user="quicksteps_tests",
        password="tests",
        database="quicksteps_tests"
    )
    cursor = db.cursor()

    process_one_direction('xls/Dveri diz zakazy (XLSX).xlsx', 'xls/Dveri diz tovary (XLSX).xlsx')
    process_one_direction('xls/Keramika diz zakazy (XLSX).xlsx', 'xls/Keramika diz tovary (XLSX).xlsx')
    process_one_direction('xls/Parket diz zakazy (XLSX).xlsx', 'xls/Parket diz tovary (XLSX).xlsx')

    process_users_vars()

    db.commit()
    db.close()

    # cursor.execute("DROP TABLE IF EXISTS orders")
    #
    # sql = '''CREATE TABLE orders(ORDER_ID INT PRIMARY KEY AUTO_INCREMENT, ORDER_NAME VARCHAR(64), ORDER_NUM CHAR(11) NOT NULL,
    #      ORDER_DATE DATE, SHIPMENT_DATE DATE, PAYMENT_DATE DATE, BONUS INT, CUSTOMER_NAME VARCHAR(64),
    #      ORDER_SUM FLOAT(14, 2), CARD_NUM INT(4), CARD_WAS_GIVEN DATE, ADDRESS VARCHAR(128), SHOP VARCHAR(128) )'''
    # cursor.execute(sql)
    #
    # cursor.execute("DROP TABLE IF EXISTS items")
    #
    # sql = '''CREATE TABLE items(ID INT PRIMARY KEY AUTO_INCREMENT, ORDER_ID INT, ORDER_NUM CHAR(11) NOT NULL, NAME VARCHAR(128), AMOUNT FLOAT(8, 3), AMOUNT_TYPE VARCHAR(10), PRICE FLOAT(14, 2) )'''
    # cursor.execute(sql)



    # cur_orders = read_deals()
    #
    # check_goods = process_orders(cur_orders)
    #
    # cur_goods = read_goods(check_goods)
    #
    # for order_goods in cur_goods.values():
    #     for good in order_goods:
    #         for i_el in range(len(good)):
    #             if good[i_el] is None:
    #                 good[i_el] = ''
    #
    #         cursor.execute("""INSERT INTO items(ORDER_ID, ORDER_NUM, NAME, AMOUNT, AMOUNT_TYPE, PRICE)
    #                     VALUES(%s, %s, %s, %s, %s, %s)""", good)